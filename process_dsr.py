"""
DSR Second Brain — Daily Processor
===================================
Run this (or have Claude run it) each morning after uploading the DSR Excel to SharePoint.

Usage:
  python process_dsr.py <path_to_excel>

What it does:
  1. Parses the DSR Excel file (same logic as the browser dashboard)
  2. Appends today's snapshot to knowledge/dsr_history.json
  3. Regenerates DSR_Dashboard.html with historical trends baked in
"""

import sys
import json
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
KNOWLEDGE    = os.path.join(BASE_DIR, "knowledge", "dsr_history.json")
TARGETS_FILE = os.path.join(BASE_DIR, "knowledge", "targets.json")
DASHBOARD    = os.path.join(BASE_DIR, "DSR_Dashboard.html")

# ── Load knowledge base ───────────────────────────────────────────────────────
def load_knowledge():
    if not os.path.exists(KNOWLEDGE):
        return {"_meta": {}, "snapshots": []}
    with open(KNOWLEDGE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_knowledge(kb):
    with open(KNOWLEDGE, "w", encoding="utf-8") as f:
        json.dump(kb, f, indent=2, ensure_ascii=False)

def load_targets():
    if not os.path.exists(TARGETS_FILE):
        return {}
    with open(TARGETS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Strip _meta key — only return month entries
    return {k: v for k, v in data.items() if not k.startswith("_")}

# ── Parse DSR Excel ───────────────────────────────────────────────────────────
def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Find summary sheet
    sheet_name = next(
        (n for n in wb.sheetnames if "summary" in n.lower()),
        wb.sheetnames[0]
    )
    ws = wb[sheet_name]

    # Read all rows into a list
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    # Auto-detect label column (same logic as JS dashboard)
    ANCHORS = ['per sales group','per project','sales channel','payment term',
               'sales target','as of','daily sales report','i. per','ii. per',
               'iv-a','vi-a','vii.']
    name_col = 1  # fallback
    for ri, row in enumerate(data[:60]):
        if not row: continue
        for ci in range(min(4, len(row))):
            cell = str(row[ci] or "").lower().strip()
            if any(a in cell for a in ANCHORS):
                name_col = ci
                break
        else:
            continue
        break

    def v(row, col):
        """Get numeric value at column index."""
        try:
            val = row[col] if col < len(row) else None
            return float(val) if isinstance(val, (int, float)) else 0.0
        except:
            return 0.0

    def s(row, col):
        """Get string value at column index."""
        try:
            val = row[col] if col < len(row) else None
            return str(val).strip() if val is not None else ""
        except:
            return ""

    def c(n):
        return n + (name_col - 1)

    def find_row(label):
        label_l = label.lower()
        for i, row in enumerate(data):
            if row and s(row, name_col).lower().find(label_l) >= 0:
                return i
        return -1

    def grand_total_after(start):
        for i in range(start + 2, len(data)):
            row = data[i]
            if row and s(row, name_col) == "Grand Total":
                return row
        return None

    snap = {
        "date": "",
        "shortDate": "",
        "processedAt": datetime.utcnow().isoformat() + "Z",
        "sourceFile": os.path.basename(filepath),
        "s2Total": {},
        "s1Total": {},
        "s4aTotal": {},
        "s7Total": {},
        "projects": [],
        "salesGroups": [],
        "etsTermsTotal": {},
        "htsTermsTotal": {},
        "pdRSTotal": {},
    }

    # Date
    date_row = next((r for r in data if r and str(r[name_col] or "").lower().startswith("as of")), None)
    if date_row:
        snap["date"] = str(date_row[name_col]).replace("As of ", "").replace("as of ", "").strip()
        import re
        m = re.search(r'(\w+ \d+)', snap["date"])
        snap["shortDate"] = m.group(1) if m else snap["date"]

    # Section II — Per Project totals
    s2 = find_row("II. Per Project")
    if s2 >= 0:
        gt = grand_total_after(s2)
        if gt:
            snap["s2Total"] = {
                "rsU": v(gt, c(3)), "rsNCP": v(gt, c(4)),
                "psSlots": v(gt, c(10)), "psNCP": v(gt, c(11))
            }
        # Project group totals
        current_group = None
        for i in range(s2 + 3, min(s2 + 35, len(data))):
            row = data[i]
            if not row or row[name_col] is None: continue
            name = s(row, name_col)
            if not name or name == "Grand Total": break
            if name in ["ETS", "HTS", "New Projects"]:
                current_group = name
                snap["projects"].append({
                    "name": name, "isGroup": True,
                    "rsU": v(row, c(3)), "rsNCP": v(row, c(4))
                })

    # Section I — Sales Groups (Grand Total only)
    s1 = find_row("I. Per Sales Group")
    if s1 >= 0:
        gt = grand_total_after(s1)
        if gt:
            snap["s1Total"] = {
                "rsU": v(gt, c(3)), "rsNCP": v(gt, c(4)),
                "totU": v(gt, c(7)), "totNCP": v(gt, c(8)),
                "psSlots": v(gt, c(10)), "psNCP": v(gt, c(11))
            }

    # Section IV-A — Channel Grand Total
    s4a = find_row("IV-A")
    if s4a >= 0:
        gt = grand_total_after(s4a)
        if gt:
            snap["s4aTotal"] = {
                "localU": v(gt, c(17)), "localNCP": v(gt, c(18)),
                "intlU": v(gt, c(19)), "intlNCP": v(gt, c(20)),
                "totU": v(gt, c(21)), "totNCP": v(gt, c(22))
            }

    # Section VI-A — ETS Grand Total
    s6a = find_row("VI-A")
    if s6a >= 0:
        gt = grand_total_after(s6a)
        if gt:
            snap["etsTermsTotal"] = {"u": v(gt, c(13)) or v(gt, c(21)), "ncp": v(gt, c(14)) or v(gt, c(22))}

    # Section VI-B — HTS Grand Total
    s6b = find_row("VI-B")
    if s6b >= 0:
        gt = grand_total_after(s6b)
        if gt:
            snap["htsTermsTotal"] = {"u": v(gt, c(19)), "ncp": v(gt, c(20))}

    # Section VII — Performance Grand Total
    s7 = find_row("VII. Sales")
    if s7 >= 0:
        gt = grand_total_after(s7)
        if gt:
            snap["s7Total"] = {
                "tgt": v(gt, c(5)), "act": v(gt, c(8)),
                "perf": v(gt, c(14))
            }

    # Section III-A — Per PD RS Total
    s3a = find_row("III-A. Per")
    if s3a >= 0:
        gt = grand_total_after(s3a)
        if gt:
            snap["pdRSTotal"] = {
                "totU": v(gt, c(23)), "totNCP": v(gt, c(24)),
                "psU": v(gt, c(26)), "psNCP": v(gt, c(27))
            }

    return snap

# ── Inject targets into dashboard HTML ───────────────────────────────────────
def inject_targets_into_dashboard(targets):
    if not os.path.exists(DASHBOARD):
        return
    with open(DASHBOARD, "r", encoding="utf-8") as f:
        html = f.read()
    import re
    targets_json = json.dumps(targets, ensure_ascii=False)
    new_block = (
        "<!--DSR_TARGETS_BLOCK_START-->\n"
        f"<script>window.DSR_TARGETS = {targets_json};</script>\n"
        "<!--DSR_TARGETS_BLOCK_END-->"
    )
    if "<!--DSR_TARGETS_BLOCK_START-->" in html:
        html = re.sub(
            r'<!--DSR_TARGETS_BLOCK_START-->.*?<!--DSR_TARGETS_BLOCK_END-->',
            new_block, html, flags=re.DOTALL
        )
        with open(DASHBOARD, "w", encoding="utf-8") as f:
            f.write(html)
        month_count = len([k for k in targets if not k.startswith("_")])
        print(f"  Targets injected — {month_count} month(s) configured.")

# ── Inject history into dashboard HTML ───────────────────────────────────────
def inject_history_into_dashboard(kb):
    if not os.path.exists(DASHBOARD):
        print(f"  Dashboard not found at {DASHBOARD} — skipping injection.")
        return

    with open(DASHBOARD, "r", encoding="utf-8") as f:
        html = f.read()

    history_json = json.dumps(kb["snapshots"], ensure_ascii=False)

    # Replace safe placeholder block (avoids corrupting </body> inside JS strings)
    import re
    new_block = (
        "<!--DSR_HISTORY_BLOCK_START-->\n"
        f"<script>window.DSR_HISTORY = {history_json};</script>\n"
        "<!--DSR_HISTORY_BLOCK_END-->"
    )

    if "<!--DSR_HISTORY_BLOCK_START-->" in html:
        updated = re.sub(
            r'<!--DSR_HISTORY_BLOCK_START-->.*?<!--DSR_HISTORY_BLOCK_END-->',
            new_block,
            html,
            flags=re.DOTALL
        )
    else:
        # Fallback: append before </body> if placeholder not found
        updated = html.replace("</body>", new_block + "\n</body>", 1)

    with open(DASHBOARD, "w", encoding="utf-8") as f:
        f.write(updated)

    print(f"  Dashboard updated with {len(kb['snapshots'])} historical snapshot(s).")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python process_dsr.py <path_to_excel>")
        print("\nNote: Claude can also run this for you — just say 'process today's DSR'.")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)

    print(f"\n📂 Processing: {os.path.basename(filepath)}")

    # Parse
    print("  Parsing Excel...")
    snap = parse_excel(filepath)
    print(f"  Date: {snap['date']}")
    print(f"  RS: {snap['s2Total'].get('rsU', 0)} units / ₱{round(snap['s2Total'].get('rsNCP', 0)/1e6)}M")

    # Load knowledge base
    kb = load_knowledge()

    # Check for duplicate (same date)
    existing_dates = [s["shortDate"] for s in kb["snapshots"]]
    if snap["shortDate"] in existing_dates:
        print(f"  ⚠️  Snapshot for {snap['shortDate']} already exists — updating it.")
        kb["snapshots"] = [s for s in kb["snapshots"] if s["shortDate"] != snap["shortDate"]]

    # Append and sort by date
    kb["snapshots"].append(snap)
    kb["snapshots"].sort(key=lambda s: s.get("processedAt", ""))

    # Save
    save_knowledge(kb)
    print(f"  ✅ Knowledge base updated — {len(kb['snapshots'])} total snapshot(s).")

    # Inject targets
    targets = load_targets()
    inject_targets_into_dashboard(targets)

    # Inject history into dashboard
    inject_history_into_dashboard(kb)

    # Also copy as index.html for GitHub Pages
    index_path = os.path.join(BASE_DIR, "index.html")
    try:
        import shutil
        shutil.copy2(DASHBOARD, index_path)
    except Exception:
        pass

    print("Done. Open DSR_Dashboard.html to view the updated dashboard.")

if __name__ == "__main__":
    main()
